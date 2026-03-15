#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
查询周缺货率数据脚本
口径与 card 1474 保持一致：
  - 数据源：smart_cooker_sg.point_session_log（展开 commodity_list JSON）
  - 过滤：commodity_type=1 的主食商品；排除 chill+ 点位（ai_device.spice_cabinet_type=5, device_type=8）
  - 满配：point_size 1/2 → 8 SKU；point_size 3 → 6 SKU
  - 缺货率 = 1 - SUM(LEAST(在线SKU数, 满配SKU数)) / SUM(满配SKU数)
  - session 数 = COUNT(DISTINCT point_id+session_id)
"""

import sys
import os

# 添加 code 目录到 Python 路径，以便导入 lib 模块
code_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../code'))
sys.path.insert(0, code_dir)

from lib import create_db_connection
import pandas as pd
from datetime import datetime


def query_weekly_shortage():
    """
    查询周缺货率数据（口径与 card 1474 一致）
    从 smart_cooker_sg.point_session_log 展开 JSON 实时计算，
    按周（周日为起始日）聚合最近 42 天的数据。

    Returns:
        list: 周缺货率数据列表
    """
    print("正在连接数据库 (smart_cooker_sg)...")
    db = create_db_connection(mysql_database='smart_cooker_sg')

    print("正在查询周缺货率数据...")

    sql = """
        WITH
          -- 1) 展开 JSON，读取最近 42 天的 session 明细
          raw_items AS (
            SELECT
              DATE(psl.create_time)  AS dt,
              psl.point_id,
              psl.session_id,
              jt.id                  AS commodity_id,
              jt.qty                 AS qty
            FROM point_session_log psl
            JOIN JSON_TABLE(
              psl.commodity_list,
              '$[*]' COLUMNS (id INT PATH '$.id', qty INT PATH '$.qty')
            ) jt
            WHERE
              psl.create_time >= '2026-01-01 00:00:00'
              AND psl.create_time >= DATE_SUB(CURDATE(), INTERVAL 42 DAY)
              AND psl.create_time <  DATE_ADD(CURDATE(), INTERVAL 1 DAY)
              AND jt.qty > 0
              -- 排除 chill+ 点位
              AND psl.point_id NOT IN (
                SELECT DISTINCT a.point_id
                FROM ai_device a
                LEFT JOIN point      p ON p.id = a.point_id
                LEFT JOIN point_ext  e ON p.id = e.point_id
                WHERE a.device_type = 8
                  AND a.point_id > 0
              )
          ),

          -- 2) 过滤：主食商品（commodity_type=1）且排除汤料柜点位
          valid_items AS (
            SELECT r.dt, r.point_id, r.session_id, r.commodity_id
            FROM raw_items r
            JOIN commodity c ON c.id = r.commodity_id AND c.commodity_type = 1
            WHERE NOT EXISTS (
              SELECT 1 FROM ai_device d
              WHERE d.point_id = r.point_id
                AND d.point_id <> 0
                AND d.spice_cabinet_type = 5
                AND d.device_type = 8
            )
          ),

          -- 3) session 粒度：统计每次点餐的在线 SKU 数
          session_online AS (
            SELECT dt, point_id, session_id,
                   COUNT(DISTINCT commodity_id) AS online_dish_cnt
            FROM valid_items
            GROUP BY dt, point_id, session_id
          ),

          -- 4) 点位满配 SKU 数（按 point_size 定义上限）
          point_full AS (
            SELECT
              pe.point_id,
              CASE
                WHEN pe.point_size IN (1, 2) THEN 8
                WHEN pe.point_size = 3       THEN 6
                ELSE NULL
              END AS full_session_dish_cnt
            FROM point_ext pe
          ),

          -- 5) session KPI：实际上架数取 LEAST，判断是否缺货
          session_kpi AS (
            SELECT
              s.dt,
              s.point_id,
              s.session_id,
              LEAST(s.online_dish_cnt, pf.full_session_dish_cnt) AS session_dish_cnt,
              pf.full_session_dish_cnt
            FROM session_online s
            JOIN point_full pf
              ON pf.point_id = s.point_id
             AND pf.full_session_dish_cnt IS NOT NULL
          ),

          -- 6) 按周聚合（周日为起始日）
          weekly_data AS (
            SELECT
              DATE_FORMAT(
                dt - INTERVAL (DAYOFWEEK(dt) - 1) DAY,
                '%Y-%m-%d'
              ) AS `周起始日(周日)`,
              ROUND(
                1 - SUM(session_dish_cnt) / SUM(full_session_dish_cnt),
                4
              ) AS `缺货率(sku权重)`,
              COUNT(DISTINCT CONCAT(point_id, '-', session_id)) AS `session数`,
              COUNT(DISTINCT dt) AS days_count
            FROM session_kpi
            GROUP BY `周起始日(周日)`
          )

        SELECT
          `周起始日(周日)`,
          `缺货率(sku权重)`,
          `session数`
        FROM weekly_data
        WHERE days_count >= 7  -- 只显示完整的周（7天数据）
        ORDER BY `周起始日(周日)` DESC
    """
    
    results = db.execute_query(sql)
    print(f"✅ 查询成功！共查询到 {len(results)} 周的数据\n")
    
    # 将缺货率转换为百分比（乘以100）
    for row in results:
        if '缺货率(sku权重)' in row and row['缺货率(sku权重)'] is not None:
            row['缺货率(sku权重)'] = round(row['缺货率(sku权重)'] * 100, 2)
    
    return results


def display_shortage_data(data):
    """
    显示周缺货率数据
    
    Args:
        data: 周缺货率数据列表
    """
    if not data:
        print("没有查询到任何数据")
        return
    
    # 转换为 DataFrame 以便更好地展示
    df = pd.DataFrame(data)
    
    print("=" * 100)
    print("周缺货率数据统计（最近42天）")
    print("=" * 100)
    print(f"\n总计周数: {len(df)}")
    
    # 显示列信息
    print(f"\n数据列: {', '.join(df.columns.tolist())}")
    
    # 显示所有数据（因为数据量不大）
    print("\n详细数据:")
    print("-" * 100)
    print(df.to_string(index=False))
    
    # 显示数据统计信息
    print("\n" + "=" * 100)
    print("数据统计摘要")
    print("=" * 100)
    
    # 缺货率统计（已经是百分比格式）
    if '缺货率(sku权重)' in df.columns:
        shortage_rate = df['缺货率(sku权重)']
        print(f"\n缺货率(sku权重)统计:")
        print(f"  平均值: {shortage_rate.mean():.2f}%")
        print(f"  最小值: {shortage_rate.min():.2f}%")
        print(f"  最大值: {shortage_rate.max():.2f}%")
        print(f"  中位数: {shortage_rate.median():.2f}%")
    
    # session数统计
    if 'session数' in df.columns:
        session_cnt = df['session数']
        print(f"\nsession数统计:")
        print(f"  总计: {session_cnt.sum():,}")
        print(f"  平均值: {session_cnt.mean():.2f}")
        print(f"  最小值: {session_cnt.min()}")
        print(f"  最大值: {session_cnt.max()}")
    
    return df


def export_to_csv(data, output_file=None):
    """
    导出数据到 CSV 文件
    
    Args:
        data: 周缺货率数据列表
        output_file: 输出文件路径（可选）
    """
    if not data:
        print("没有数据可以导出")
        return
    
    # 如果没有指定输出文件，使用默认文件名
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'weekly_shortage_{timestamp}.csv'
    
    # 转换为 DataFrame
    df = pd.DataFrame(data)
    
    # 导出到 CSV
    output_path = os.path.join(os.path.dirname(__file__), output_file)
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    
    print(f"\n✅ 数据已导出到: {output_path}")
    print(f"   文件大小: {os.path.getsize(output_path) / 1024:.2f} KB")


def export_to_excel(data, output_file=None):
    """
    导出数据到 Excel 文件（带格式）
    
    Args:
        data: 周缺货率数据列表
        output_file: 输出文件路径（可选）
    """
    if not data:
        print("没有数据可以导出")
        return
    
    try:
        # 如果没有指定输出文件，使用默认文件名
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f'周缺货率_{timestamp}.xlsx'
        
        # 转换为 DataFrame
        df = pd.DataFrame(data)
        
        # 确保数值列是数值类型
        if '缺货率(sku权重)' in df.columns:
            df['缺货率(sku权重)'] = pd.to_numeric(df['缺货率(sku权重)'], errors='coerce')
        if 'session数' in df.columns:
            df['session数'] = pd.to_numeric(df['session数'], errors='coerce')
        
        # 导出到 Excel
        output_path = os.path.join(os.path.dirname(__file__), output_file)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='周缺货率', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['周缺货率']
            
            # 导入样式模块
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 设置列宽
            column_widths = {
                'A': 18,  # 周起始日(周日)
                'B': 20,  # 缺货率(sku权重)
                'C': 15,  # session数
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置数据行格式
            for row_idx in range(2, len(df) + 2):
                # 周起始日 - 居中对齐
                worksheet[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # 缺货率 - 百分比格式（保留2位小数，带%符号）
                worksheet[f'B{row_idx}'].number_format = '0.00"%"'
                worksheet[f'B{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
                
                # session数 - 数值格式（千位分隔符）
                worksheet[f'C{row_idx}'].number_format = '#,##0'
                worksheet[f'C{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=3):
                for cell in row:
                    cell.border = thin_border
        
        print(f"\n✅ 数据已导出到 Excel: {output_path}")
        print(f"   文件大小: {os.path.getsize(output_path) / 1024:.2f} KB")
        
    except Exception as e:
        print(f"\n⚠️  Excel 导出失败: {str(e)}")
        print("   已自动导出为 CSV 格式")


def main():
    """
    主函数
    """
    print("=" * 100)
    print("周缺货率数据查询工具")
    print("=" * 100)
    print()
    
    try:
        # 查询数据
        data = query_weekly_shortage()
        
        # 显示数据
        df = display_shortage_data(data)
        
        
        # 导出数据到 Excel
        export_to_excel(data)
        
        print("\n" + "=" * 100)
        print("✅ 所有操作完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

