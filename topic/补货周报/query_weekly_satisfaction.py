#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按周统计满足率
周日到周六为一周
"""

import sys
import os
from datetime import datetime, timedelta
from collections import defaultdict

# 添加 code 目录到 Python 路径，以便导入 lib 模块
code_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../code'))
sys.path.insert(0, code_dir)

from lib import create_db_connection
import pandas as pd


def query_fenjian_log():
    """
    查询分拣日志数据
    
    Returns:
        list: 分拣日志数据列表
    """
    print("正在连接数据库 (smart_cooker_sg)...")
    db = create_db_connection(mysql_database='smart_cooker_sg')
    
    print("正在查询分拣日志数据（最近90天）...")
    sql = """
        SELECT
          a.id,
          a.req_task_id,
          a.req_car_id,
          a.create_time,
          s.veg_box_count as 旧补货数,
          a.point_forecast_amount,
          a.point_remain_stock,
          CASE 
            WHEN (a.point_forecast_amount - a.point_remain_stock) > 0 
            THEN ROUND(s.veg_box_count / (a.point_forecast_amount - a.point_remain_stock), 4)
            ELSE NULL 
          END as 满足率
        FROM
          point_commodity_fenjian_log a
          LEFT JOIN sorting_tasks s ON a.req_task_id = s.task_id
        WHERE
          a.status = 1 
          AND a.req_task_id != 0
          AND a.create_time >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
        ORDER BY
          a.create_time DESC
    """
    
    results = db.execute_query(sql)
    print(f"✅ 查询成功！共查询到 {len(results)} 条记录\n")
    
    return results


def deduplicate_by_req_task_id(data):
    """
    按 req_task_id 去重，相同 req_task_id 只保留一条 id 最大的数据
    
    Args:
        data: 原始数据列表
        
    Returns:
        list: 去重后的数据列表
    """
    print("正在对 req_task_id 进行去重...")
    
    # 按 req_task_id 分组，保留 id 最大的记录
    task_id_map = {}
    for row in data:
        task_id = row['req_task_id']
        row_id = row['id']
        
        if task_id not in task_id_map:
            task_id_map[task_id] = row
        else:
            # 如果当前记录的 id 更大，则替换
            if row_id > task_id_map[task_id]['id']:
                task_id_map[task_id] = row
    
    deduplicated_data = list(task_id_map.values())
    print(f"去重前: {len(data)} 条记录")
    print(f"去重后: {len(deduplicated_data)} 条记录\n")
    
    return deduplicated_data


def get_week_start(date):
    """
    获取指定日期所在周的周日日期
    
    Args:
        date: 日期对象
        
    Returns:
        date: 周日日期
    """
    # weekday() 返回 0-6，0是周一，6是周日
    # 我们需要将周日作为一周的开始
    weekday = date.weekday()
    if weekday == 6:  # 如果是周日
        return date
    else:
        # 计算到上一个周日的天数
        days_since_sunday = (weekday + 1) % 7
        return date - timedelta(days=days_since_sunday)


def calculate_weekly_satisfaction(data):
    """
    按周统计平均满足率
    
    Args:
        data: 去重后的数据列表
        
    Returns:
        list: 按周统计的结果
    """
    print("正在按周统计满足率...")
    
    # 正常车的 req_car_id 列表
    NORMAL_CAR_IDS = [2, 14, 15]
    
    # 按周分组统计
    weekly_stats = defaultdict(lambda: {
        'all_satisfaction': [],
        'normal_satisfaction': [],
        'virtual_satisfaction': []
    })
    
    for row in data:
        # 获取满足率
        satisfaction = row.get('满足率')
        if satisfaction is None:
            continue
        
        # 获取日期
        create_time = row.get('create_time')
        if isinstance(create_time, str):
            date = datetime.strptime(create_time, '%Y-%m-%d %H:%M:%S').date()
        else:
            date = create_time.date() if hasattr(create_time, 'date') else create_time
        
        # 获取周日日期
        week_start = get_week_start(date)
        week_key = week_start.strftime('%Y/%m/%d')
        
        # 获取车辆类型
        req_car_id = row.get('req_car_id')
        
        # 添加到全部统计
        weekly_stats[week_key]['all_satisfaction'].append(satisfaction)
        
        # 根据车辆类型分类统计
        if req_car_id in NORMAL_CAR_IDS:
            weekly_stats[week_key]['normal_satisfaction'].append(satisfaction)
        else:
            weekly_stats[week_key]['virtual_satisfaction'].append(satisfaction)
    
    # 计算平均值并格式化结果
    results = []
    for week_key in sorted(weekly_stats.keys(), reverse=True):
        stats = weekly_stats[week_key]
        
        # 计算全部平均满足率
        all_avg = sum(stats['all_satisfaction']) / len(stats['all_satisfaction']) if stats['all_satisfaction'] else 0
        
        # 计算正常车平均满足率
        normal_avg = sum(stats['normal_satisfaction']) / len(stats['normal_satisfaction']) if stats['normal_satisfaction'] else 0
        
        # 计算虚拟车平均满足率
        virtual_avg = sum(stats['virtual_satisfaction']) / len(stats['virtual_satisfaction']) if stats['virtual_satisfaction'] else 0
        
        results.append({
            '日期': week_key,
            '全部': round(all_avg * 100, 2),  # 转换为百分比
            '正常车': round(normal_avg * 100, 2),
            '虚拟车': round(virtual_avg * 100, 2)
        })
    
    print(f"✅ 统计完成！共 {len(results)} 周的数据\n")
    
    return results


def export_to_excel(data, output_file=None):
    """
    导出数据到 Excel 文件
    
    Args:
        data: 统计结果数据列表
        output_file: 输出文件路径（可选）
    """
    if not data:
        print("没有数据可以导出")
        return
    
    try:
        # 如果没有指定输出文件，使用默认文件名
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f'周满足率_{timestamp}.xlsx'
        
        # 转换为 DataFrame
        df = pd.DataFrame(data)
        
        # 确保数值列是数值类型
        for col in ['全部', '正常车', '虚拟车']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 导出到 Excel
        output_path = os.path.join(os.path.dirname(__file__), output_file)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='周满足率', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['周满足率']
            
            # 导入样式模块
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 设置列宽
            column_widths = {
                'A': 15,  # 日期
                'B': 12,  # 全部
                'C': 12,  # 正常车
                'D': 12,  # 虚拟车
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置数据行格式
            for row_idx in range(2, len(df) + 2):
                # 日期 - 居中对齐
                worksheet[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # 全部 - 百分比格式（保留2位小数，带%符号）
                worksheet[f'B{row_idx}'].number_format = '0.00"%"'
                worksheet[f'B{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
                
                # 正常车 - 百分比格式（保留2位小数，带%符号）
                worksheet[f'C{row_idx}'].number_format = '0.00"%"'
                worksheet[f'C{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
                
                # 虚拟车 - 百分比格式（保留2位小数，带%符号）
                worksheet[f'D{row_idx}'].number_format = '0.00"%"'
                worksheet[f'D{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
        
        print(f"\n✅ 数据已导出到 Excel: {output_path}")
        print(f"   文件大小: {os.path.getsize(output_path) / 1024:.2f} KB")
        
    except Exception as e:
        print(f"\n⚠️  Excel 导出失败: {str(e)}")
        import traceback
        traceback.print_exc()


def main():
    """
    主函数
    """
    print("=" * 100)
    print("周满足率统计工具")
    print("=" * 100)
    print()
    
    try:
        # 1. 查询数据
        data = query_fenjian_log()
        
        # 2. 去重
        deduplicated_data = deduplicate_by_req_task_id(data)
        
        # 3. 按周统计满足率
        weekly_results = calculate_weekly_satisfaction(deduplicated_data)
        
        # 4. 显示结果
        if weekly_results:
            df = pd.DataFrame(weekly_results)
            print("=" * 100)
            print("周满足率统计结果（最近90天）")
            print("=" * 100)
            print(f"\n总计周数: {len(df)}\n")
            print(df.to_string(index=False))
        
        # 5. 导出数据到 Excel
        export_to_excel(weekly_results)
        
        print("\n" + "=" * 100)
        print("✅ 所有操作完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

