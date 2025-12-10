#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
周日均补货数统计脚本
按周统计补货数据（周日到周六为一周）
"""

import sys
import os
import numpy as np

# 添加 code 目录到 Python 路径，以便导入 lib 模块
code_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../code'))
sys.path.insert(0, code_dir)

from lib import create_db_connection
import pandas as pd
from datetime import datetime
from decimal import Decimal


def query_weekly_replenishment():
    """
    查询周补货统计数据
    
    Returns:
        list: 周补货统计数据列表
    """
    print("正在连接数据库 (smart_cooker_sg)...")
    # 使用 smart_cooker_sg 数据库
    db = create_db_connection(mysql_database='smart_cooker_sg')
    
    print("正在查询周补货统计数据（最近42天）...")
    sql = """
        WITH daily_stats AS (
          SELECT
            DATE(sorting_start_time) AS '日期',
            SUM(veg_box_count) AS '日补货数',
            COUNT( point_id) AS '日点位数'
          FROM sorting_tasks
          WHERE sorting_start_time >= DATE_SUB(CURDATE(), INTERVAL 42 DAY)
          GROUP BY DATE(sorting_start_time)
        ),
        weekly_stats AS (
          SELECT
            -- 周标识：取该周的周日（格式：YYYY/M/D），作为每周的唯一标识
            DATE_FORMAT(
              STR_TO_DATE(`日期`, '%Y-%m-%d') - INTERVAL (DAYOFWEEK(STR_TO_DATE(`日期`, '%Y-%m-%d')) - 1) DAY,
              '%Y/%c/%e'
            ) AS '周数',
            -- 补货数：周总补货数
            SUM(`日补货数`) AS '补货数',
            -- 日均补货：周总补货数 / 6
            ROUND(SUM(`日补货数`) / 6, 0) AS '日均补货',
            -- 日最大补货数：该周内单日最大补货数
            MAX(`日补货数`) AS '日最大补货数',
            -- 日均点位数：周平均点位数
            ROUND(SUM(`日点位数`) / 6, 0) AS '日均点位',
            -- 用于排序的周起始日期
            MIN(STR_TO_DATE(`日期`, '%Y-%m-%d') - INTERVAL (DAYOFWEEK(STR_TO_DATE(`日期`, '%Y-%m-%d')) - 1) DAY) AS week_start,
            -- 统计该周有几天的数据
            COUNT(DISTINCT `日期`) AS days_count
          FROM daily_stats
          GROUP BY
            DATE_FORMAT(
              STR_TO_DATE(`日期`, '%Y-%m-%d') - INTERVAL (DAYOFWEEK(STR_TO_DATE(`日期`, '%Y-%m-%d')) - 1) DAY,
              '%Y/%c/%e'
            )
        )
        SELECT
          `周数`,
          `补货数`,
          `日均补货`,
          `日最大补货数`,
          `日均点位`
        FROM weekly_stats
        WHERE days_count >= 7  -- 只显示完整的周（7天数据）
        ORDER BY week_start DESC
    """
    
    results = db.execute_query(sql)
    print(f"✅ 查询成功！共查询到 {len(results)} 周的数据\n")
    
    # 移除辅助列 week_start
    for row in results:
        if 'week_start' in row:
            del row['week_start']
    
    return results


def convert_to_python_types(data):
    """
    将数据中的 numpy 类型和 Decimal 类型转换为 Python 原生类型
    
    Args:
        data: 数据列表（字典格式）
    
    Returns:
        list: 转换后的数据列表
    """
    converted_data = []
    for row in data:
        converted_row = {}
        for key, value in row.items():
            # 确保键是字符串类型
            str_key = str(key) if not isinstance(key, str) else key
            
            # 处理 None 值
            if value is None:
                converted_row[str_key] = None
            # 处理 numpy 数组类型
            elif isinstance(value, np.ndarray):
                converted_row[str_key] = value.tolist()
            # 处理 numpy 标量类型（包括所有子类型）
            elif isinstance(value, (np.integer, np.floating, np.bool_, np.complexfloating, np.generic)):
                converted_row[str_key] = value.item()
            # 处理 Decimal 类型
            elif isinstance(value, Decimal):
                converted_row[str_key] = float(value)
            # 处理 datetime 类型（保持原样，pandas 可以处理）
            elif isinstance(value, (datetime, pd.Timestamp)):
                converted_row[str_key] = value
            # 处理 bytes 类型
            elif isinstance(value, bytes):
                converted_row[str_key] = value.decode('utf-8', errors='ignore')
            # 其他类型保持原样
            else:
                converted_row[str_key] = value
        converted_data.append(converted_row)
    return converted_data


def display_weekly_data(data):
    """
    显示周补货数据
    
    Args:
        data: 周补货数据列表
    """
    if not data:
        print("没有查询到任何数据")
        return
    
    # 将数据转换为 Python 原生类型
    converted_data = convert_to_python_types(data)
    
    # 转换为 DataFrame 以便更好地展示
    df = pd.DataFrame(converted_data)
    
    print("=" * 100)
    print("周日均补货数统计（最近42天）")
    print("=" * 100)
    print(f"\n总计周数: {len(df)}")
    
    # 显示列信息
    print(f"\n数据列: {', '.join(df.columns.tolist())}")
    
    # 显示所有数据
    print("\n详细数据:")
    print("-" * 100)
    print(df.to_string(index=False))
    
    # 显示数据统计信息
    print("\n" + "=" * 100)
    print("数据统计摘要")
    print("=" * 100)
    
    # 补货数统计
    if '补货数' in df.columns:
        print(f"\n补货数统计:")
        print(f"  总计: {df['补货数'].sum():,.0f} 盒")
        print(f"  周平均: {df['补货数'].mean():,.0f} 盒")
        print(f"  最小值: {df['补货数'].min():,.0f} 盒")
        print(f"  最大值: {df['补货数'].max():,.0f} 盒")
    
    # 日均补货统计
    if '日均补货' in df.columns:
        print(f"\n日均补货统计:")
        print(f"  平均值: {df['日均补货'].mean():,.0f} 盒/天")
        print(f"  最小值: {df['日均补货'].min():,.0f} 盒/天")
        print(f"  最大值: {df['日均补货'].max():,.0f} 盒/天")
    
    # 日最大补货数统计
    if '日最大补货数' in df.columns:
        print(f"\n日最大补货数统计:")
        print(f"  平均值: {df['日最大补货数'].mean():,.0f} 盒")
        print(f"  最小值: {df['日最大补货数'].min():,.0f} 盒")
        print(f"  最大值: {df['日最大补货数'].max():,.0f} 盒")
    
    # 日均点位数统计
    if '日均点位' in df.columns:
        print(f"\n日均点位数统计:")
        print(f"  平均值: {df['日均点位'].mean():.0f} 个")
        print(f"  最小值: {df['日均点位'].min():.0f} 个")
        print(f"  最大值: {df['日均点位'].max():.0f} 个")
    
    return df


def export_to_csv(data, output_file=None):
    """
    导出数据到 CSV 文件
    
    Args:
        data: 周补货数据列表
        output_file: 输出文件路径（可选）
    """
    if not data:
        print("没有数据可以导出")
        return
    
    # 如果没有指定输出文件，使用默认文件名
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'weekly_replenishment_{timestamp}.csv'
    
    # 将数据转换为 Python 原生类型
    converted_data = convert_to_python_types(data)
    
    # 转换为 DataFrame
    df = pd.DataFrame(converted_data)
    
    # 导出到 CSV
    output_path = os.path.join(os.path.dirname(__file__), output_file)
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    
    print(f"\n✅ 数据已导出到: {output_path}")
    print(f"   文件大小: {os.path.getsize(output_path) / 1024:.2f} KB")


def export_to_excel(data, output_file=None):
    """
    导出数据到 Excel 文件（带格式）
    
    Args:
        data: 周补货数据列表
        output_file: 输出文件路径（可选）
    """
    if not data:
        print("没有数据可以导出")
        return
    
    try:
        # 如果没有指定输出文件，使用默认文件名
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f'周日均补货数_{timestamp}.xlsx'
        
        # 将数据转换为 Python 原生类型
        converted_data = convert_to_python_types(data)
        
        # 转换为 DataFrame
        df = pd.DataFrame(converted_data)
        
        # 确保数值列是数值类型（而不是字符串）
        numeric_columns = ['补货数', '日均补货', '日最大补货数', '日均点位']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 导出到 Excel
        output_path = os.path.join(os.path.dirname(__file__), output_file)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='周日均补货数', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['周日均补货数']
            
            # 导入样式模块
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 设置列宽
            column_widths = {
                'A': 15,  # 周数
                'B': 12,  # 补货数
                'C': 12,  # 日均补货
                'D': 15,  # 日最大补货数
                'E': 12,  # 日均点位
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置数据行格式
            for row_idx in range(2, len(df) + 2):
                # 周数列 - 居中对齐
                worksheet[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # 补货数 - 数值格式（千位分隔符）
                worksheet[f'B{row_idx}'].number_format = '#,##0'
                worksheet[f'B{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
                
                # 日均补货 - 数值格式（千位分隔符）
                worksheet[f'C{row_idx}'].number_format = '#,##0'
                worksheet[f'C{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
                
                # 日最大补货数 - 数值格式（千位分隔符）
                worksheet[f'D{row_idx}'].number_format = '#,##0'
                worksheet[f'D{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
                
                # 日均点位 - 数值格式（整数）
                worksheet[f'E{row_idx}'].number_format = '0'
                worksheet[f'E{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=5):
                for cell in row:
                    cell.border = thin_border
        
        print(f"\n✅ 数据已导出到 Excel: {output_path}")
        print(f"   文件大小: {os.path.getsize(output_path) / 1024:.2f} KB")
        
    except Exception as e:
        print(f"\n⚠️  Excel 导出失败: {str(e)}")
        print("   已自动导出为 CSV 格式")


def main():
    """
    主函数
    """
    print("=" * 100)
    print("周日均补货数统计工具")
    print("=" * 100)
    print()
    
    try:
        # 查询数据
        data = query_weekly_replenishment()
        
        # 显示数据
        df = display_weekly_data(data)
      
        
        # 尝试导出到 Excel
        export_to_excel(data)
        
        print("\n" + "=" * 100)
        print("✅ 所有操作完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

