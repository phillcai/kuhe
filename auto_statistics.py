#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动统计系统 - 实时更新自评估Excel文件的统计数据
支持数据变化检测和自动重新计算
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
import os
import time
from datetime import datetime
import hashlib

class AutoStatisticsSystem:
    """
    自动统计系统类
    """
    
    def __init__(self, excel_path):
        """
        初始化自动统计系统
        
        Args:
            excel_path: Excel文件路径
        """
        self.excel_path = excel_path
        self.last_hash = None
        self.talent_types = [
            'W - Wonder（洞察）',
            'I - Invention（创意）',
            'D - Discernment（判断）',
            'G - Galvanizing（动员）',
            'E - Enablement（使能）',
            'T - Tenacity（交付）'
        ]
    
    def calculate_file_hash(self):
        """
        计算文件哈希值，用于检测文件是否被修改
        
        Returns:
            str: 文件哈希值
        """
        try:
            with open(self.excel_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except Exception as e:
            print(f"计算文件哈希时出错: {e}")
            return None
    
    def read_and_process_data(self):
        """
        读取并处理Excel数据
        
        Returns:
            pd.DataFrame: 处理后的数据
        """
        try:
            df = pd.read_excel(self.excel_path)
            
            # 向前填充天赋类型
            df['天赋类型'] = df['天赋类型'].ffill()
            
            # 过滤有效数据
            df_filtered = df[
                (~df['天赋类型'].str.contains('使用说明|1\.|2\.|3\.', na=False)) &
                (df['评估条目'].notna()) &  # 确保评估条目不为空
                (df['天赋类型'].isin(self.talent_types))  # 确保是有效的天赋类型
            ]
            
            # 手动计算得分（因为Excel公式可能没有生效）
            def calculate_score(row):
                if pd.isna(row['得分']):
                    # 手动计算得分
                    if row['经常（2 分）'] == '是':
                        return 2
                    elif row['偶尔（1 分）'] == '是':
                        return 1
                    elif row['很少（0 分）'] == '是':
                        return 0
                    else:
                        return 0
                else:
                    return row['得分']
            
            df_filtered = df_filtered.copy()  # 避免警告
            df_filtered['得分'] = df_filtered.apply(calculate_score, axis=1)
            
            return df_filtered
            
        except Exception as e:
            print(f"读取数据时出错: {e}")
            return None
    
    def calculate_statistics(self, df):
        """
        计算统计数据
        
        Args:
            df: 处理后的数据框
            
        Returns:
            list: 统计数据列表
        """
        statistics_data = []
        
        for talent_type in self.talent_types:
            talent_rows = df[df['天赋类型'] == talent_type]
            
            if len(talent_rows) > 0:
                scores = talent_rows['得分'].fillna(0)
                total_score = scores.sum()
                max_possible_score = len(talent_rows) * 2
                avg_score = scores.mean()
                question_count = len(talent_rows)
                score_rate = (total_score / max_possible_score * 100) if max_possible_score > 0 else 0
                
                statistics_data.append({
                    '天赋类型': talent_type,
                    '题目数量': question_count,
                    '总得分': total_score,
                    '平均得分': round(avg_score, 2),
                    '最高可能得分': max_possible_score,
                    '得分率(%)': round(score_rate, 1)
                })
        
        return statistics_data
    
    def update_statistics_sheet(self, statistics_data):
        """
        更新统计工作表
        
        Args:
            statistics_data: 统计数据列表
        """
        try:
            wb = load_workbook(self.excel_path)
            
            # 删除旧的统计工作表
            if '天赋类型统计' in wb.sheetnames:
                del wb['天赋类型统计']
            
            stats_ws = wb.create_sheet('天赋类型统计')
            
            # 设置标题
            stats_ws['A1'] = f'天赋类型得分统计表（自动更新） - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            stats_ws['A1'].font = Font(size=16, bold=True)
            stats_ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            stats_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            
            # 设置表头
            headers = ['天赋类型', '题目数量', '总得分', '平均得分', '最高可能得分', '得分率(%)']
            for col, header in enumerate(headers, 1):
                cell = stats_ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 填入数据
            for row, data in enumerate(statistics_data, 4):
                stats_ws.cell(row=row, column=1, value=data['天赋类型'])
                stats_ws.cell(row=row, column=2, value=data['题目数量'])
                stats_ws.cell(row=row, column=3, value=data['总得分'])
                stats_ws.cell(row=row, column=4, value=data['平均得分'])
                stats_ws.cell(row=row, column=5, value=data['最高可能得分'])
                stats_ws.cell(row=row, column=6, value=data['得分率(%)'])
                
                # 设置数据行样式
                for col in range(1, 7):
                    cell = stats_ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # 设置列宽
            column_widths = [25, 10, 10, 10, 12, 12]
            for col, width in enumerate(column_widths, 1):
                stats_ws.column_dimensions[chr(64 + col)].width = width
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(3, len(statistics_data) + 4):
                for col in range(1, 7):
                    stats_ws.cell(row=row, column=col).border = thin_border
            
            # 添加总计行
            total_row = len(statistics_data) + 5
            stats_ws.cell(row=total_row, column=1, value='总计')
            stats_ws.cell(row=total_row, column=2, value=sum(data['题目数量'] for data in statistics_data))
            stats_ws.cell(row=total_row, column=3, value=sum(data['总得分'] for data in statistics_data))
            avg_score = sum(data['平均得分'] for data in statistics_data) / len(statistics_data) if len(statistics_data) > 0 else 0
            stats_ws.cell(row=total_row, column=4, value=round(avg_score, 2))
            stats_ws.cell(row=total_row, column=5, value=sum(data['最高可能得分'] for data in statistics_data))
            total_possible = sum(data['最高可能得分'] for data in statistics_data)
            overall_rate = (sum(data['总得分'] for data in statistics_data) / total_possible * 100) if total_possible > 0 else 0
            stats_ws.cell(row=total_row, column=6, value=round(overall_rate, 1))
            
            # 设置总计行样式
            for col in range(1, 7):
                cell = stats_ws.cell(row=total_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # 创建得分率图表
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "各天赋类型得分率对比（自动更新）"
            chart.y_axis.title = '得分率(%)'
            chart.x_axis.title = '天赋类型'
            
            # 设置图表数据
            data = Reference(stats_ws, min_col=6, min_row=3, max_row=len(statistics_data) + 3, max_col=6)
            categories = Reference(stats_ws, min_col=1, min_row=4, max_row=len(statistics_data) + 3, max_col=1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # 设置图表位置
            stats_ws.add_chart(chart, "H3")
            
            # 添加说明文字
            note_row = len(statistics_data) + 7
            stats_ws.cell(row=note_row, column=1, value="说明:")
            stats_ws.cell(row=note_row + 1, column=1, value="1. 此表格会自动检测数据变化并更新")
            stats_ws.cell(row=note_row + 2, column=1, value="2. 每个天赋类型包含3个题目，总共18题")
            stats_ws.cell(row=note_row + 3, column=1, value="3. 得分率 = 总得分 / 最高可能得分 × 100%")
            stats_ws.cell(row=note_row + 4, column=1, value="4. 平均得分 = 该天赋类型所有题目得分的平均值")
            
            # 设置说明文字样式
            note_font = Font(size=10, italic=True, color="666666")
            for i in range(note_row, note_row + 5):
                stats_ws.cell(row=i, column=1).font = note_font
            
            # 保存文件
            wb.save(self.excel_path)
            print(f"统计表格已更新: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
        except Exception as e:
            print(f"更新统计表格时出错: {e}")
    
    def add_auto_calculation_formulas(self):
        """
        添加自动计算公式到主工作表
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # 找到得分列（最后一列）
            score_column = ws.max_column
            
            # 为每一行添加自动计算公式
            for row in range(2, ws.max_row + 1):
                # 检查是否是有效的数据行
                if ws.cell(row=row, column=2).value and not str(ws.cell(row=row, column=2).value).startswith('使用说明'):
                    # 添加得分计算公式
                    score_formula = f'=IF(C{row}="是",2,IF(D{row}="是",1,IF(E{row}="是",0,0)))'
                    ws.cell(row=row, column=score_column, value=score_formula)
            
            wb.save(self.excel_path)
            print("自动计算公式已添加")
            
        except Exception as e:
            print(f"添加自动计算公式时出错: {e}")
    
    def monitor_and_update(self, check_interval=5):
        """
        监控文件变化并自动更新统计
        
        Args:
            check_interval: 检查间隔（秒）
        """
        print(f"开始监控文件: {self.excel_path}")
        print(f"检查间隔: {check_interval}秒")
        print("按 Ctrl+C 停止监控")
        
        try:
            while True:
                current_hash = self.calculate_file_hash()
                
                if current_hash != self.last_hash:
                    print(f"\n检测到文件变化: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    
                    # 读取并处理数据
                    df = self.read_and_process_data()
                    if df is not None:
                        # 计算统计数据
                        statistics_data = self.calculate_statistics(df)
                        
                        # 更新统计表格
                        self.update_statistics_sheet(statistics_data)
                        
                        # 显示统计结果
                        self.display_statistics(statistics_data)
                    
                    self.last_hash = current_hash
                
                time.sleep(check_interval)
                
        except KeyboardInterrupt:
            print("\n监控已停止")
        except Exception as e:
            print(f"监控过程中出错: {e}")
    
    def display_statistics(self, statistics_data):
        """
        显示统计数据
        
        Args:
            statistics_data: 统计数据列表
        """
        print("\n天赋类型统计结果:")
        print("-" * 80)
        for data in statistics_data:
            print(f"{data['天赋类型']:<20} | 题目:{data['题目数量']:>2} | 得分:{data['总得分']:>2}/{data['最高可能得分']:>2} | 得分率:{data['得分率(%)']:>5}%")
        print("-" * 80)
        
        # 计算总体统计
        total_questions = sum(data['题目数量'] for data in statistics_data)
        total_score = sum(data['总得分'] for data in statistics_data)
        max_possible = sum(data['最高可能得分'] for data in statistics_data)
        overall_rate = (total_score / max_possible * 100) if max_possible > 0 else 0
        
        print(f"总体统计: 题目{total_questions}题 | 得分{total_score}/{max_possible} | 得分率{overall_rate:.1f}%")
    
    def run_once(self):
        """
        运行一次统计更新
        """
        print("执行一次统计更新...")
        
        # 读取并处理数据
        df = self.read_and_process_data()
        if df is not None:
            # 计算统计数据
            statistics_data = self.calculate_statistics(df)
            
            # 更新统计表格
            self.update_statistics_sheet(statistics_data)
            
            # 显示统计结果
            self.display_statistics(statistics_data)
        else:
            print("无法读取数据")

def main():
    """
    主函数
    """
    excel_path = "/Users/admin/Code/MyCode/kuhe/data/自评估_修改版.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        return
    
    # 创建自动统计系统
    auto_stats = AutoStatisticsSystem(excel_path)
    
    print("自动统计系统")
    print("1. 运行一次统计更新")
    print("2. 开始监控模式（自动检测文件变化）")
    print("3. 添加自动计算公式")
    
    choice = input("请选择操作 (1/2/3): ").strip()
    
    if choice == "1":
        auto_stats.run_once()
    elif choice == "2":
        interval = input("请输入检查间隔（秒，默认5秒）: ").strip()
        try:
            interval = int(interval) if interval else 5
        except ValueError:
            interval = 5
        auto_stats.monitor_and_update(interval)
    elif choice == "3":
        auto_stats.add_auto_calculation_formulas()
    else:
        print("无效选择")

if __name__ == "__main__":
    main()
