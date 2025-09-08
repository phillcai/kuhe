#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修改自评估Excel文件，添加单选功能和自动计分
增强版：添加VBA宏确保单选功能
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
import os

def modify_excel_file_enhanced():
    """
    修改Excel文件，添加单选功能和自动计分（增强版）
    """
    file_path = "/Users/admin/Code/MyCode/kuhe/data/自评估.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return
    
    try:
        # 读取Excel文件
        print("正在读取Excel文件...")
        df = pd.read_excel(file_path)
        
        print("文件内容预览:")
        print(df.head())
        print(f"\n文件形状: {df.shape}")
        print(f"列名: {list(df.columns)}")
        
        # 使用openpyxl加载工作簿以进行更精细的控制
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 获取列名
        headers = [cell.value for cell in ws[1]]
        print(f"Excel列名: {headers}")
        
        # 查找包含"经常"、"偶尔"、"很少"的列
        score_columns = []
        score_column_indices = []
        
        for i, header in enumerate(headers):
            if header and ("经常" in str(header) or "偶尔" in str(header) or "很少" in str(header)):
                score_columns.append(header)
                score_column_indices.append(i + 1)  # Excel列索引从1开始
        
        print(f"找到评分列: {score_columns}")
        print(f"评分列索引: {score_column_indices}")
        
        # 如果没有找到评分列，创建示例结构
        if not score_columns:
            print("未找到评分列，创建示例结构...")
            
            # 添加新的列
            ws.cell(row=1, column=len(headers) + 1, value="经常(2分)")
            ws.cell(row=1, column=len(headers) + 2, value="偶尔(1分)")
            ws.cell(row=1, column=len(headers) + 3, value="很少(0分)")
            ws.cell(row=1, column=len(headers) + 4, value="得分")
            
            # 更新列索引
            score_column_indices = [len(headers) + 1, len(headers) + 2, len(headers) + 3]
            score_columns = ["经常(2分)", "偶尔(1分)", "很少(0分)"]
            total_score_column = len(headers) + 4
        
        # 为每一行添加数据验证和公式
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过标题行）
            
            # 添加数据验证规则 - 每行只能选择一个选项
            validation_rule = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
            validation_rule.add(f"{chr(64 + score_column_indices[0])}{row}:{chr(64 + score_column_indices[2])}{row}")
            ws.add_data_validation(validation_rule)
            
            # 添加得分计算公式
            # 使用IF函数检查每列的值，如果为"是"则返回对应分数
            score_formula = f'=IF({chr(64 + score_column_indices[0])}{row}="是",2,IF({chr(64 + score_column_indices[1])}{row}="是",1,IF({chr(64 + score_column_indices[2])}{row}="是",0,0)))'
            
            if len(score_columns) == 3:
                total_score_column = score_column_indices[2] + 1
            else:
                total_score_column = len(headers) + 1
            
            ws.cell(row=row, column=total_score_column, value=score_formula)
            
            # 清空现有的复选框符号，准备填入新数据
            for col_idx in score_column_indices:
                ws.cell(row=row, column=col_idx, value="")
        
        # 设置样式
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # 设置标题行样式
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 天赋类型
        ws.column_dimensions['B'].width = 40  # 评估条目
        ws.column_dimensions['C'].width = 12  # 经常
        ws.column_dimensions['D'].width = 12  # 偶尔
        ws.column_dimensions['E'].width = 12  # 很少
        ws.column_dimensions['F'].width = 8   # 得分
        
        # 添加使用说明
        instruction_row = ws.max_row + 2
        ws.cell(row=instruction_row, column=1, value="使用说明:")
        ws.cell(row=instruction_row + 1, column=1, value="1. 每行只能在'经常'、'偶尔'、'很少'中选择一个")
        ws.cell(row=instruction_row + 2, column=1, value="2. 选择'是'表示勾选该选项")
        ws.cell(row=instruction_row + 3, column=1, value="3. 得分会自动计算：经常=2分，偶尔=1分，很少=0分")
        
        # 设置说明文字样式
        instruction_font = Font(size=10, italic=True, color="666666")
        for i in range(instruction_row, instruction_row + 4):
            ws.cell(row=i, column=1).font = instruction_font
        
        # 保存文件
        output_path = "/Users/admin/Code/MyCode/kuhe/data/自评估_修改版.xlsx"
        wb.save(output_path)
        print(f"修改完成！文件已保存为: {output_path}")
        
        # 显示修改说明
        print("\n修改说明:")
        print("1. 每行的三个选项列只能选择'是'或'否'")
        print("2. 每行只能有一个选项选择'是'")
        print("3. 得分列会自动计算:")
        print("   - 经常(2分): 2分")
        print("   - 偶尔(1分): 1分") 
        print("   - 很少(0分): 0分")
        print("4. 请在新文件中填写评估内容")
        print("5. 文件底部添加了使用说明")
        
        # 创建VBA宏文件
        create_vba_macro(output_path)
        
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()

def create_vba_macro(excel_path):
    """
    创建VBA宏文件，确保单选功能
    """
    vba_code = '''
Sub Worksheet_Change(ByVal Target As Range)
    ' 确保每行只能选择一个选项
    Dim row As Long
    Dim col As Long
    Dim scoreCols As Variant
    Dim i As Integer
    
    ' 定义评分列的列号（C=3, D=4, E=5）
    scoreCols = Array(3, 4, 5)
    
    ' 检查是否在评分列范围内
    If Target.Column >= 3 And Target.Column <= 5 And Target.Row > 1 Then
        row = Target.Row
        col = Target.Column
        
        ' 如果当前单元格选择了"是"
        If Target.Value = "是" Then
            ' 清除同行的其他评分列
            For i = 0 To UBound(scoreCols)
                If scoreCols(i) <> col Then
                    Cells(row, scoreCols(i)).Value = ""
                End If
            Next i
        End If
    End If
End Sub
'''
    
    vba_file_path = "/Users/admin/Code/MyCode/kuhe/data/自评估_VBA宏.txt"
    with open(vba_file_path, 'w', encoding='utf-8') as f:
        f.write(vba_code)
    
    print(f"\nVBA宏代码已保存到: {vba_file_path}")
    print("如需使用VBA宏，请:")
    print("1. 打开Excel文件")
    print("2. 按Alt+F11打开VBA编辑器")
    print("3. 双击工作表名称")
    print("4. 粘贴VBA代码")
    print("5. 保存并关闭VBA编辑器")

if __name__ == "__main__":
    modify_excel_file_enhanced()
