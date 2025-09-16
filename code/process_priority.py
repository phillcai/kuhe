#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
优先级数据处理脚本
功能：处理data/priority.csv文件，为每个req_id内的点位按priority降序排列，添加sort字段
"""

import pandas as pd
import os
from pathlib import Path


def process_priority_data():
    """
    处理优先级数据的主函数
    
    功能说明：
    1. 读取data/priority.csv文件
    2. 按req_id分组，在每组内按priority降序排列
    3. 为每组添加sort字段，最高priority的sort=1，依次递增
    4. 保存处理后的数据到新文件
    """
    
    # 设置文件路径
    input_file = Path("data/priority.csv")
    output_file = Path("data/priority_with_sort.csv")
    
    try:
        # 检查输入文件是否存在
        if not input_file.exists():
            print(f"错误：输入文件 {input_file} 不存在！")
            return False
        
        print(f"正在读取文件：{input_file}")
        # 读取CSV文件
        df = pd.read_csv(input_file)
        
        print(f"原始数据形状：{df.shape}")
        print(f"包含的req_id数量：{df['req_id'].nunique()}")
        
        # 按req_id分组，在每组内按priority降序排列，然后添加sort字段
        print("正在处理数据...")
        df_sorted = df.sort_values(['req_id', 'priority'], ascending=[True, False])
        
        # 为每个req_id组内添加sort字段（从1开始）
        df_sorted['sort'] = df_sorted.groupby('req_id').cumcount() + 1
        
        # 新增priority_new字段
        # 当sku_cnt<8且point_stock-sales_score>15时，priority_new为priority/1.5，否则为priority
        condition = (df_sorted['sku_cnt'] < 8) & (df_sorted['point_stock'] - df_sorted['sales_score'] > 15)
        df_sorted['priority_new'] = df_sorted['priority'].where(~condition, df_sorted['priority'] / 1.5)
        
        # 按req_id分组，根据priority_new降序排列，添加sort_new字段
        df_sorted = df_sorted.groupby('req_id').apply(
            lambda x: x.sort_values('priority_new', ascending=False).reset_index(drop=True)
        ).reset_index(drop=True)
        
        # 为每个req_id组内添加sort_new字段（从1开始）
        df_sorted['sort_new'] = df_sorted.groupby('req_id').cumcount() + 1
        
        # 验证处理结果
        print("\n处理结果验证：")
        for req_id in df_sorted['req_id'].unique()[:3]:  # 只显示前3个req_id的结果
            req_data = df_sorted[df_sorted['req_id'] == req_id][['req_id', 'point_id', 'priority', 'priority_new', 'sort', 'sort_new']].head()
            print(f"\nreq_id: {req_id}")
            print(req_data.to_string(index=False))
        
        # 保存处理后的数据
        print(f"\n正在保存处理后的数据到：{output_file}")
        df_sorted.to_csv(output_file, index=False, encoding='utf-8')
        
        # 创建带条件格式的Excel文件
        excel_output_file = Path("data/priority_with_sort_highlighted.xlsx")
        print(f"正在创建带高亮的Excel文件：{excel_output_file}")
        
        # 找出需要高亮的行：sort <= 8 且 sort_new > 8
        highlight_condition = (df_sorted['sort'] <= 8) & (df_sorted['sort_new'] > 8)
        highlight_rows = df_sorted[highlight_condition]
        
        print(f"找到 {len(highlight_rows)} 行符合高亮条件 (sort <= 8 且 sort_new > 8)")
        
        if len(highlight_rows) > 0:
            print("符合条件的点位详情：")
            for _, row in highlight_rows.iterrows():
                print(f"  req_id: {row['req_id']}, point_id: {row['point_id']}, sort: {int(row['sort'])} → sort_new: {int(row['sort_new'])}")
        
        # 创建Excel文件并应用条件格式
        try:
            with pd.ExcelWriter(excel_output_file, engine='openpyxl') as writer:
                df_sorted.to_excel(writer, sheet_name='优先级数据', index=False)
                
                # 获取工作表
                worksheet = writer.sheets['优先级数据']
                
                # 导入openpyxl的样式模块
                from openpyxl.styles import PatternFill
                
                # 定义黄色填充
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                
                # 为符合条件的行应用黄色背景
                for idx in highlight_rows.index:
                    excel_row = idx + 2  # Excel行号从2开始（1是标题行）
                    for col in range(1, len(df_sorted.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = yellow_fill
            
            print(f"✅ Excel文件已创建：{excel_output_file}")
        except ImportError:
            print("⚠️  openpyxl未安装，无法创建Excel文件。请运行: uv add openpyxl")
        except Exception as e:
            print(f"❌ 创建Excel文件时发生错误：{str(e)}")
        
        print(f"✅ 处理完成！")
        print(f"📊 处理统计：")
        print(f"   - 总记录数：{len(df_sorted)}")
        print(f"   - req_id数量：{df_sorted['req_id'].nunique()}")
        print(f"   - CSV输出文件：{output_file}")
        print(f"   - Excel输出文件：{excel_output_file}")
        print(f"   - 高亮行数：{len(highlight_rows)}")
        
        return True
        
    except Exception as e:
        print(f"❌ 处理过程中发生错误：{str(e)}")
        return False


def show_sample_results():
    """
    显示处理结果的示例数据
    """
    output_file = Path("data/priority_with_sort.csv")
    
    if not output_file.exists():
        print("输出文件不存在，请先运行处理函数。")
        return
    
    try:
        df = pd.read_csv(output_file)
        
        print("\n📋 处理结果示例：")
        print("=" * 80)
        
        # 显示每个req_id的前几条记录
        for req_id in df['req_id'].unique()[:2]:  # 显示前2个req_id
            req_data = df[df['req_id'] == req_id]
            print(f"\n🔍 req_id: {req_id} (共{len(req_data)}个点位)")
            print("-" * 60)
            sample_data = req_data[['point_id', 'priority', 'sort', '点位名称']].head(10)
            print(sample_data.to_string(index=False))
            
    except Exception as e:
        print(f"❌ 显示结果时发生错误：{str(e)}")


if __name__ == "__main__":
    print("🚀 开始处理优先级数据...")
    print("=" * 60)
    
    # 执行数据处理
    success = process_priority_data()
    
    if success:
        # 显示处理结果示例
        show_sample_results()
    else:
        print("❌ 数据处理失败，请检查错误信息。")